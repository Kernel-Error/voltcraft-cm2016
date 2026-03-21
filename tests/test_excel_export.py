"""Tests for cm2016.export.excel_export — Excel/Spreadsheet export."""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import load_workbook

from cm2016.export.excel_export import export_excel
from cm2016.protocol import parse_frame
from cm2016.session import SlotRecord
from tests.conftest import make_frame

if TYPE_CHECKING:
    from pathlib import Path


def _make_records(count: int = 5) -> list[SlotRecord]:
    frame = parse_frame(
        make_frame(
            slot_overrides={
                0: {
                    "active": 1,
                    "program": 1,
                    "step": 1,
                    "status": 0x00,
                    "runtime_minutes": 10,
                    "voltage_mv": 1320,
                    "current_raw": 500,
                    "charge_cap_raw": 75000,
                },
            }
        )
    )
    return [SlotRecord.from_slot_data(frame.slots[0], "NiMH") for _ in range(count)]


class TestExcelExport:
    """Test Excel/Spreadsheet export."""

    def test_creates_valid_xlsx(self, tmp_path: Path) -> None:
        path = tmp_path / "test.xlsx"
        count = export_excel(_make_records(3), path, slot_name="Slot 1")
        assert count == 3
        assert path.exists()

        wb = load_workbook(str(path))
        ws = wb.active
        assert ws is not None
        assert ws.title == "Slot 1"

    def test_correct_headers(self, tmp_path: Path) -> None:
        path = tmp_path / "test.xlsx"
        export_excel(_make_records(1), path)

        wb = load_workbook(str(path))
        ws = wb.active
        assert ws is not None
        headers = [cell.value for cell in ws[1]]
        assert "Slot" in headers
        assert "Voltage (V)" in headers
        assert "Chemistry" in headers

    def test_correct_data_values(self, tmp_path: Path) -> None:
        path = tmp_path / "test.xlsx"
        export_excel(_make_records(1), path)

        wb = load_workbook(str(path))
        ws = wb.active
        assert ws is not None
        row = [cell.value for cell in ws[2]]
        assert row[0] == 1  # Slot
        assert row[4] == 1.32  # Voltage
        assert row[5] == 0.5  # Current

    def test_has_charts(self, tmp_path: Path) -> None:
        path = tmp_path / "test.xlsx"
        export_excel(_make_records(5), path)

        wb = load_workbook(str(path))
        ws = wb.active
        assert ws is not None
        assert len(ws._charts) == 2  # Voltage + Current charts

    def test_empty_records(self, tmp_path: Path) -> None:
        path = tmp_path / "empty.xlsx"
        count = export_excel([], path)
        assert count == 0
        assert path.exists()
