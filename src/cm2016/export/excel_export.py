"""Excel export for CM2016 recording data with embedded charts."""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from pathlib import Path

    from cm2016.session import SlotRecord

logger = logging.getLogger(__name__)

HEADERS = [
    "Slot",
    "Time",
    "Program",
    "Actual",
    "Voltage (V)",
    "Current (A)",
    "CCAP (mAh)",
    "DCAP (mAh)",
    "Chemistry",
]


def export_excel(records: list[SlotRecord], path: Path, slot_name: str = "") -> int:
    """Export slot records to an Excel .xlsx file with embedded charts.

    Args:
        records: List of SlotRecord objects to export.
        path: Output file path.
        slot_name: Name for the sheet title (e.g., "Slot 1").

    Returns:
        Number of rows written (excluding header).
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = slot_name or "Data"

    # Header row
    ws.append(HEADERS)

    # Data rows
    for record in records:
        ws.append(
            [
                record.slot_index + 1,
                record.runtime_formatted,
                record.program,
                record.status,
                round(record.voltage, 3),
                round(record.current, 3),
                round(record.charge_capacity, 2),
                round(record.discharge_capacity, 2),
                record.chemistry,
            ]
        )

    if records:
        data_end = len(records) + 1  # +1 for header

        # Voltage chart
        v_chart = LineChart()
        v_chart.title = "Voltage"
        v_chart.y_axis.title = "Voltage (V)"
        v_chart.x_axis.title = "Time"
        v_chart.style = 10
        v_chart.width = 20
        v_chart.height = 12

        v_data = Reference(ws, min_col=5, min_row=1, max_row=data_end)
        v_chart.add_data(v_data, titles_from_data=True)

        # Color voltage series green
        if v_chart.series:
            v_chart.series[0].graphicalProperties.line.solidFill = "4E9A06"

        ws.add_chart(v_chart, f"{get_column_letter(11)}2")

        # Current chart
        i_chart = LineChart()
        i_chart.title = "Current"
        i_chart.y_axis.title = "Current (A)"
        i_chart.x_axis.title = "Time"
        i_chart.style = 10
        i_chart.width = 20
        i_chart.height = 12

        i_data = Reference(ws, min_col=6, min_row=1, max_row=data_end)
        i_chart.add_data(i_data, titles_from_data=True)

        # Color current series
        if i_chart.series:
            i_chart.series[0].graphicalProperties.line.solidFill = "CC0000"

        ws.add_chart(i_chart, f"{get_column_letter(11)}18")

    # Auto-size columns
    for col_idx, header in enumerate(HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header) + 2, 12)

    wb.save(str(path))
    logger.info("Exported %d rows to %s", len(records), path)
    return len(records)
